from __future__ import annotations

import httpx

from surveyer.config import SnowballConfig
from surveyer.models import Record
from surveyer.snowball import OpenAlexSnowball
from surveyer.sources.base import HttpClient


def _work(wid, title, refs=None):
    return {
        "id": f"https://openalex.org/{wid}",
        "title": title,
        "referenced_works": [f"https://openalex.org/{r}" for r in (refs or [])],
    }


def _handler(request: httpx.Request) -> httpx.Response:
    path = request.url.path
    params = request.url.params
    if "/doi:" in path:
        # seed lookup
        return httpx.Response(200, json=_work("W1", "Seed", refs=["W2", "W3"]))
    flt = params.get("filter", "")
    if flt.startswith("openalex_id:"):  # backward batch
        return httpx.Response(
            200,
            json={
                "meta": {"count": 2},
                "results": [
                    {"title": "Ref two", "doi": "https://doi.org/10.1/two"},
                    {"title": "Ref three", "doi": "https://doi.org/10.1/three"},
                ],
            },
        )
    if flt.startswith("cites:"):  # forward
        return httpx.Response(
            200,
            json={
                "meta": {"count": 1},
                "results": [{"title": "Citing paper", "doi": "https://doi.org/10.1/c"}],
            },
        )
    return httpx.Response(404, json={})


def _client(tmp_path):
    return HttpClient(cache_dir=tmp_path, transport=httpx.MockTransport(_handler))


def test_fetch_both_directions(tmp_path):
    sb = OpenAlexSnowball(_client(tmp_path))
    seeds = [Record(title="Seed", doi="10.1/seed")]
    result = sb.fetch(seeds, SnowballConfig(enabled=True, direction="both"))
    assert result.seeds_total == 1
    assert result.seeds_resolved == 1
    assert result.backward == 2
    assert result.forward == 1
    assert len(result.candidates) == 3
    # provenance is tagged so PRISMA/export can tell the arm + direction apart
    titles = {r.title: r.query_labels for r in result.candidates}
    assert titles["Ref two"] == ["snowball:backward"]
    assert titles["Citing paper"] == ["snowball:forward"]
    assert all(r.sources == ["openalex"] for r in result.candidates)


def test_fetch_backward_only(tmp_path):
    sb = OpenAlexSnowball(_client(tmp_path))
    seeds = [Record(title="Seed", doi="10.1/seed")]
    result = sb.fetch(seeds, SnowballConfig(enabled=True, direction="backward"))
    assert result.backward == 2
    assert result.forward == 0


def test_seed_without_doi_is_skipped(tmp_path):
    sb = OpenAlexSnowball(_client(tmp_path))
    seeds = [Record(title="No DOI seed")]  # doi is None
    result = sb.fetch(seeds, SnowballConfig(enabled=True, direction="both"))
    assert result.seeds_total == 1
    assert result.seeds_resolved == 0
    assert result.candidates == []


def test_forward_is_capped(tmp_path):
    def many(request: httpx.Request) -> httpx.Response:
        path = request.url.path
        params = request.url.params
        if "/doi:" in path:
            return httpx.Response(200, json=_work("W1", "Seed"))
        if params.get("filter", "").startswith("cites:"):
            page = int(params.get("page", "1"))
            per = int(params["per-page"])
            start = (page - 1) * per
            results = [{"title": f"C{i}"} for i in range(start, start + per)]
            return httpx.Response(200, json={"meta": {"count": 9999}, "results": results})
        return httpx.Response(200, json={"meta": {"count": 0}, "results": []})

    client = HttpClient(cache_dir=tmp_path, transport=httpx.MockTransport(many))
    sb = OpenAlexSnowball(client)
    seeds = [Record(title="Seed", doi="10.1/seed")]
    result = sb.fetch(
        seeds, SnowballConfig(enabled=True, direction="forward", max_results_per_seed=5)
    )
    assert result.forward == 5


def test_snowball_stage_dedups_and_screens(tmp_path):
    from surveyer.config import (
        FilterConfig,
        KeywordConfig,
        LLMConfig,
        ProjectConfig,
        SearchConfig,
        SnowballConfig,
        SurveyConfig,
    )
    from surveyer.snowball import OpenAlexSnowball, snowball_stage

    # Candidates: a graph paper (keep), an off-topic paper (keyword-drop),
    # and one that duplicates an already-seen record (dedup-drop).
    def handler(request: httpx.Request) -> httpx.Response:
        if "/doi:" in request.url.path:
            return httpx.Response(200, json=_work("W1", "Seed", refs=["W2", "W3", "W4"]))
        flt = request.url.params.get("filter", "")
        if flt.startswith("openalex_id:"):
            return httpx.Response(
                200,
                json={
                    "meta": {"count": 3},
                    "results": [
                        {
                            "title": "Graph neural networks survey",
                            "abstract_inverted_index": {"graph": [0]},
                        },
                        {"title": "Cooking pasta recipes", "abstract_inverted_index": {"food": [0]}},
                        {"title": "Already seen paper", "doi": "https://doi.org/10.1/seen"},
                    ],
                },
            )
        return httpx.Response(200, json={"meta": {"count": 0}, "results": []})

    client = HttpClient(cache_dir=tmp_path, transport=httpx.MockTransport(handler))
    fetcher = OpenAlexSnowball(client)

    cfg = SurveyConfig(
        project=ProjectConfig(name="t", output_dir=str(tmp_path)),
        search=SearchConfig(sources=["openalex"], queries=[]),
        filter=FilterConfig(
            keyword=KeywordConfig(include=["graph"], exclude=[]),
            llm=LLMConfig(enabled=True, threshold=0.5, survey_abstract="s"),
        ),
        snowball=SnowballConfig(enabled=True, direction="backward"),
    )

    class FakeScorer:
        def score(self, survey_abstract, record, *, concepts=None):
            return (0.9, "ok") if "graph" in (record.abstract or "") else (0.1, "no")

    already = [Record(title="Already seen paper", doi="10.1/seen")]
    seeds = [Record(title="Seed", doi="10.1/seed")]
    kept, excluded, led = snowball_stage(
        seeds, already, cfg, FakeScorer(), fetcher
    )

    assert [r.title for r in kept] == ["Graph neural networks survey"]
    assert led.seeds == 1
    assert led.backward == 3
    assert led.identified == 3
    assert led.duplicates_removed == 1  # "Already seen paper"
    assert led.excluded_keyword == 1  # "Cooking pasta recipes"
    assert led.included == 1
    assert any(r.exclusion_reason for r in excluded)


def test_run_snowball_appends_to_workbook(tmp_path):
    from openpyxl import load_workbook

    from surveyer.config import (
        FilterConfig,
        KeywordConfig,
        LLMConfig,
        ProjectConfig,
        SearchConfig,
        SnowballConfig,
        SurveyConfig,
    )
    from surveyer.export import export_xlsx
    from surveyer.models import Ledger, Record
    from surveyer.snowball import SnowballFetch, run_snowball

    # Build a screened baseline workbook with one included paper.
    baseline_path = tmp_path / "survey.xlsx"
    export_xlsx(
        [Record(title="Seed paper", doi="10.1/seed", abstract="graph")],
        [],
        Ledger(included=1),
        baseline_path,
    )

    out_dir = tmp_path / "out"
    cfg = SurveyConfig(
        project=ProjectConfig(name="t", output_dir=str(out_dir)),
        search=SearchConfig(sources=["openalex"], queries=[]),
        filter=FilterConfig(
            keyword=KeywordConfig(include=["graph"], exclude=[]),
            llm=LLMConfig(enabled=True, threshold=0.5, survey_abstract="s"),
        ),
        snowball=SnowballConfig(enabled=True, direction="backward"),
    )

    class FakeSnowball:
        def fetch(self, seeds, sb_cfg, *, cancel=None):
            return SnowballFetch(
                candidates=[Record(title="Graph follow-up", abstract="graph")],
                seeds_total=len(seeds),
                seeds_resolved=len(seeds),
                backward=1,
                forward=0,
                retrieval=[],
            )

    class FakeScorer:
        def score(self, survey_abstract, record, *, concepts=None):
            return (0.9, "ok")

    result = run_snowball(
        cfg,
        baseline_path,
        scorer=FakeScorer(),
        fetcher=FakeSnowball(),
        resolve_bibtex=False,
    )
    assert result.ledger.previously_included == 1
    assert result.ledger.snowball is not None
    assert result.ledger.snowball.included == 1
    assert result.ledger.total_included() == 2

    wb = load_workbook(out_dir / "survey.xlsx")
    titles = [row[0] for row in wb["papers"].iter_rows(min_row=2, values_only=True)]
    assert "Seed paper" in titles
    assert "Graph follow-up" in titles
