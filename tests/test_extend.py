from __future__ import annotations

import pytest
from openpyxl import Workbook, load_workbook

from surveyer.export import export_xlsx
from surveyer.extend import load_baseline, split_already_screened, Baseline
from surveyer.models import Ledger, Record


def _screened_workbook(path):
    """A survey.xlsx as the tool writes it, ready for manual edits."""
    kept = [
        Record(
            title="Kept paper",
            doi="10.1/kept",
            authors=["Ada Lovelace", "Alan Turing"],
            year=2024,
            keywords=["graphs", "ml"],
            sources=["dblp"],
            llm_score=0.91,
            bibtex="@article{kept}",
            bibtex_source="dblp",
        )
    ]
    excluded = [
        Record(title="Dropped paper", doi="10.1/drop", exclusion_reason="no graph")
    ]
    export_xlsx(kept, excluded, Ledger(included=1), path)


def test_load_baseline_roundtrip(tmp_path):
    path = tmp_path / "v1.xlsx"
    _screened_workbook(path)

    base = load_baseline(path)

    assert [r.title for r in base.included] == ["Kept paper"]
    assert base.included[0].doi == "10.1/kept"
    assert base.included[0].authors == ["Ada Lovelace", "Alan Turing"]
    assert base.included[0].year == 2024
    assert base.included[0].keywords == ["graphs", "ml"]
    assert base.included[0].llm_score == 0.91
    assert base.included[0].bibtex == "@article{kept}"
    assert [r.title for r in base.excluded] == ["Dropped paper"]
    assert base.excluded[0].exclusion_reason == "no graph"


def test_load_baseline_manual_row_needs_only_title(tmp_path):
    path = tmp_path / "v1.xlsx"
    _screened_workbook(path)
    wb = load_workbook(path)
    ws = wb["papers"]
    ws.cell(row=ws.max_row + 1, column=1, value="Hand-added paper")
    wb.save(path)

    base = load_baseline(path)

    titles = [r.title for r in base.included]
    assert "Hand-added paper" in titles
    manual = next(r for r in base.included if r.title == "Hand-added paper")
    assert manual.doi is None
    assert manual.authors == []


def test_load_baseline_unknown_columns_ignored(tmp_path):
    path = tmp_path / "v1.xlsx"
    _screened_workbook(path)
    wb = load_workbook(path)
    ws = wb["papers"]
    ws.cell(row=1, column=ws.max_column + 1, value="my_notes")
    ws.cell(row=2, column=ws.max_column, value="re-read this one")
    wb.save(path)

    base = load_baseline(path)
    assert [r.title for r in base.included] == ["Kept paper"]


def test_load_baseline_missing_sheet_errors(tmp_path):
    path = tmp_path / "v1.xlsx"
    wb = Workbook()
    wb.active.title = "papers"
    wb["papers"].append(["title"])
    wb.save(path)

    with pytest.raises(ValueError, match="excluded"):
        load_baseline(path)


def test_load_baseline_empty_title_errors(tmp_path):
    path = tmp_path / "v1.xlsx"
    _screened_workbook(path)
    wb = load_workbook(path)
    ws = wb["papers"]
    # A row with data but no title (column A is `title`).
    ws.cell(row=ws.max_row + 1, column=2, value="Anonymous Author")
    wb.save(path)

    with pytest.raises(ValueError, match="title"):
        load_baseline(path)


def test_load_baseline_missing_file_errors(tmp_path):
    with pytest.raises(ValueError, match="not found"):
        load_baseline(tmp_path / "nope.xlsx")


def test_load_baseline_tolerates_sloppy_list_separators(tmp_path):
    path = tmp_path / "v1.xlsx"
    _screened_workbook(path)
    wb = load_workbook(path)
    ws = wb["papers"]
    header = [c.value for c in ws[1]]
    authors_col = header.index("authors") + 1
    ws.cell(row=2, column=authors_col, value="Ada Lovelace;Alan Turing ;  Grace Hopper")
    wb.save(path)

    base = load_baseline(path)
    assert base.included[0].authors == ["Ada Lovelace", "Alan Turing", "Grace Hopper"]


def test_load_baseline_invalid_number_names_location(tmp_path):
    path = tmp_path / "v1.xlsx"
    _screened_workbook(path)
    wb = load_workbook(path)
    ws = wb["papers"]
    header = [c.value for c in ws[1]]
    year_col = header.index("year") + 1
    ws.cell(row=2, column=year_col, value="circa 2020")
    wb.save(path)

    with pytest.raises(ValueError, match="row 2.*year"):
        load_baseline(path)


def _baseline():
    return Baseline(
        included=[Record(title="Graph learning survey", doi="10.1/in")],
        excluded=[Record(title="Cooking with graphs", doi="10.1/out")],
    )


def test_split_drops_doi_match_from_either_sheet(tmp_path):
    fetched = [
        Record(title="Different title same doi", doi="10.1/IN "),  # case/space
        Record(title="Also refound", doi="10.1/out"),
        Record(title="Genuinely new paper", doi="10.1/new"),
    ]
    fresh, dropped = split_already_screened(fetched, _baseline())
    assert [r.title for r in fresh] == ["Genuinely new paper"]
    assert dropped == 2


def test_split_drops_fuzzy_title_match():
    fetched = [
        Record(title="Graph Learning: A Survey"),  # fuzzy match, no doi
        Record(title="Completely unrelated quantum paper"),
    ]
    fresh, dropped = split_already_screened(fetched, _baseline())
    assert [r.title for r in fresh] == ["Completely unrelated quantum paper"]
    assert dropped == 1


def test_split_keeps_everything_with_empty_baseline():
    fetched = [Record(title="A"), Record(title="B")]
    fresh, dropped = split_already_screened(fetched, Baseline(included=[], excluded=[]))
    assert len(fresh) == 2
    assert dropped == 0


def test_split_keeps_similar_title_with_conflicting_strong_doi():
    baseline = Baseline(
        included=[Record(title="Graph learning survey", doi="10.1/in")],
        excluded=[],
    )
    fetched = [Record(title="Graph learning survey", doi="10.99/other")]
    fresh, dropped = split_already_screened(fetched, baseline)
    assert [r.title for r in fresh] == ["Graph learning survey"]
    assert dropped == 0


def test_split_drops_similar_title_when_new_doi_is_weak_preprint():
    baseline = Baseline(
        included=[Record(title="Graph learning survey", doi="10.1/in")],
        excluded=[],
    )
    fetched = [Record(title="Graph learning survey", doi="10.48550/arXiv.1")]
    fresh, dropped = split_already_screened(fetched, baseline)
    assert fresh == []
    assert dropped == 1
