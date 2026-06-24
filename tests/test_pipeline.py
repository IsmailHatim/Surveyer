from __future__ import annotations

import threading

import pytest
from openpyxl import load_workbook

from surveyer.cancel import PipelineCancelled
from surveyer.config import (
    ExtendConfig,
    FilterConfig,
    KeywordConfig,
    LLMConfig,
    ProjectConfig,
    Query,
    SearchConfig,
    SurveyConfig,
)
from surveyer.export import export_xlsx
from surveyer.ledger import load_ledger
from surveyer.models import Ledger as LedgerModel
from surveyer.models import Record, SearchResult
from surveyer.pipeline import run_pipeline


class FakeSource:
    name = "fake"

    def search(self, terms, *, max_results):
        return SearchResult(
            records=[
                Record(title="Relevant security paper", abstract="relevant", doi="10.1/a"),
                Record(title="Relevant security paper", abstract="relevant", doi="10.1/a"),
                Record(title="Off topic cooking", abstract="recipes", doi="10.1/b"),
            ]
        )


class FakeScorer:
    def score(self, survey_abstract, record, *, concepts=None):
        return (0.9, "ok", {}) if "relevant" in (record.abstract or "") else (0.1, "no", {})


def _cfg(tmp_path) -> SurveyConfig:
    return SurveyConfig(
        project=ProjectConfig(name="t", output_dir=str(tmp_path)),
        search=SearchConfig(sources=["fake"], queries=[Query(label="A", terms="x")]),
        filter=FilterConfig(
            keyword=KeywordConfig(include=["security"], exclude=[]),
            llm=LLMConfig(enabled=True, threshold=0.5, survey_abstract="s"),
        ),
    )


class FakeResolver:
    def __init__(self):
        self.called_with = None

    def resolve_all(self, records):
        self.called_with = records
        for r in records:
            r.bibtex = f"@misc{{{r.title}}}"
            r.bibtex_source = "local"


def test_run_pipeline_resolves_bibtex(tmp_path):
    cfg = _cfg(tmp_path)
    resolver = FakeResolver()
    result = run_pipeline(
        cfg,
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolver=resolver,
    )
    assert resolver.called_with is result.kept
    assert all(r.bibtex for r in result.kept)
    assert (tmp_path / "references.bib").exists()


def test_run_pipeline_can_skip_bibtex(tmp_path):
    cfg = _cfg(tmp_path)
    resolver = FakeResolver()
    result = run_pipeline(
        cfg,
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolver=resolver,
        resolve_bibtex=False,
    )
    assert resolver.called_with is None  # resolver never invoked
    assert all(r.bibtex is None for r in result.kept)


def test_run_pipeline_end_to_end(tmp_path):
    cfg = _cfg(tmp_path)
    result = run_pipeline(
        cfg,
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolve_bibtex=False,  # predates bibtex feature; avoid network calls
    )
    assert result.ledger.total_identified() == 3
    assert result.ledger.duplicates_removed == 1
    assert result.ledger.excluded_keyword == 1
    assert result.ledger.included == 1
    assert (tmp_path / "survey.xlsx").exists()
    assert (tmp_path / "prisma.mmd").exists()  # always written
    led = load_ledger(tmp_path / "ledger.json")
    assert led.included == 1
    assert result.ledger.excluded_llm == 0
    assert len(result.excluded) == 1


def test_run_pipeline_forwards_refresh_to_builders(tmp_path, monkeypatch):
    import surveyer.bibtex as bibtex
    import surveyer.pipeline as pipeline

    seen: dict[str, bool] = {}

    def fake_build_registry(search, cache_root, *, refresh=False):
        seen["registry"] = refresh
        return {"fake": FakeSource()}

    def fake_build_resolver(cache_root, *, refresh=False):
        seen["resolver"] = refresh
        return FakeResolver()

    monkeypatch.setattr(pipeline, "build_registry", fake_build_registry)
    monkeypatch.setattr(bibtex, "build_resolver", fake_build_resolver)

    run_pipeline(_cfg(tmp_path), scorer=FakeScorer(), refresh=True)
    assert seen == {"registry": True, "resolver": True}


def test_run_pipeline_defaults_refresh_off(tmp_path, monkeypatch):
    import surveyer.pipeline as pipeline

    seen: dict[str, bool] = {}

    def fake_build_registry(search, cache_root, *, refresh=False):
        seen["registry"] = refresh
        return {"fake": FakeSource()}

    monkeypatch.setattr(pipeline, "build_registry", fake_build_registry)
    run_pipeline(_cfg(tmp_path), scorer=FakeScorer(), resolve_bibtex=False)
    assert seen["registry"] is False


def _baseline_xlsx(path):
    """A screened v1: one manually kept paper, one manually excluded paper.

    The excluded paper shares its DOI (10.1/b) with FakeSource's off-topic
    record, so the new run re-finds an already-screened paper.
    """
    kept = [
        Record(
            title="Old manually kept paper",
            doi="10.1/old",
            bibtex="@misc{old}",
            bibtex_source="dblp",
        )
    ]
    excluded = [
        Record(title="Off topic cooking", doi="10.1/b", exclusion_reason="manual")
    ]
    export_xlsx(kept, excluded, LedgerModel(included=1), path)


def _extend_cfg(tmp_path) -> SurveyConfig:
    baseline = tmp_path / "v1.xlsx"
    _baseline_xlsx(baseline)
    cfg = _cfg(tmp_path / "v2")
    cfg.extend = ExtendConfig(xlsx=str(baseline))
    return cfg


def test_extend_pipeline_pins_baseline_and_skips_screened(tmp_path):
    cfg = _extend_cfg(tmp_path)
    result = run_pipeline(
        cfg,
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolve_bibtex=False,
    )
    led = result.ledger
    # FakeSource: 2x "Relevant security paper" (10.1/a) + "Off topic cooking"
    # (10.1/b). 10.1/b matches the baseline's excluded sheet.
    assert led.duplicates_removed == 1
    assert led.already_screened == 1
    assert led.previously_included == 1
    assert led.included == 1  # only the genuinely new relevant paper
    assert led.total_included() == 2

    titles = {r.title for r in result.kept}
    assert titles == {"Old manually kept paper", "Relevant security paper"}
    assert all(r.title != "Off topic cooking" for r in result.kept)
    # The manually excluded paper stays in the excluded list.
    assert "Off topic cooking" in {r.title for r in result.excluded}


def test_extend_pipeline_never_rescores_baseline(tmp_path):
    class RefindingSource:
        name = "fake"

        def search(self, terms, *, max_results):
            return SearchResult(
                records=[
                    Record(
                        title="Fresh security paper", abstract="relevant", doi="10.1/new"
                    ),
                    # Same DOI as the baseline's kept paper: already screened.
                    Record(title="Old security paper", abstract="relevant", doi="10.1/old"),
                ]
            )

    class RecordingScorer:
        def __init__(self):
            self.seen = []

        def score(self, survey_abstract, record, *, concepts=None):
            self.seen.append(record.title)
            return (0.9, "ok")

    baseline = tmp_path / "v1.xlsx"
    export_xlsx(
        [Record(title="Old security paper", doi="10.1/old", bibtex="@misc{old}")],
        [],
        LedgerModel(included=1),
        baseline,
    )
    cfg = _cfg(tmp_path / "v2")
    cfg.extend = ExtendConfig(xlsx=str(baseline))

    scorer = RecordingScorer()
    result = run_pipeline(
        cfg,
        registry={"fake": RefindingSource()},
        scorer=scorer,
        resolve_bibtex=False,
    )

    # The re-found paper passes the keyword filter, so only the
    # already-screened drop keeps it away from the scorer.
    assert scorer.seen == ["Fresh security paper"]
    assert result.ledger.already_screened == 1


def test_extend_pipeline_writes_appended_workbook(tmp_path):
    cfg = _extend_cfg(tmp_path)
    run_pipeline(
        cfg,
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolve_bibtex=False,
    )
    wb = load_workbook(tmp_path / "v2" / "survey.xlsx")
    papers = [row[0].value for row in wb["papers"].iter_rows(min_row=2)]
    assert papers == ["Old manually kept paper", "Relevant security paper"]


def test_extend_pipeline_resolves_bibtex_only_where_missing(tmp_path):
    cfg = _extend_cfg(tmp_path)
    resolver = FakeResolver()
    result = run_pipeline(
        cfg,
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolver=resolver,
    )
    resolved_titles = [r.title for r in resolver.called_with]
    # The new paper needs bibtex; the baseline paper already has one.
    assert "Relevant security paper" in resolved_titles
    assert "Old manually kept paper" not in resolved_titles
    assert all(r.bibtex for r in result.kept)


def test_run_pipeline_cancel_before_work(tmp_path):
    cfg = _cfg(tmp_path)
    event = threading.Event()
    event.set()
    with pytest.raises(PipelineCancelled):
        run_pipeline(
            cfg,
            registry={"fake": FakeSource()},
            scorer=FakeScorer(),
            resolve_bibtex=False,
            cancel=event,
        )
    assert not (tmp_path / "ledger.json").exists()
    assert not (tmp_path / "survey.xlsx").exists()
    assert not (tmp_path / "prisma.svg").exists()


def test_run_pipeline_cancel_during_llm(tmp_path):
    event = threading.Event()

    class CancelScorer:
        def score(self, survey_abstract, record, *, concepts=None):
            event.set()  # trip cancel as soon as the LLM stage starts
            return 0.9, "ok"

    cfg = _cfg(tmp_path)
    with pytest.raises(PipelineCancelled):
        run_pipeline(
            cfg,
            registry={"fake": FakeSource()},
            scorer=CancelScorer(),
            resolve_bibtex=False,
            cancel=event,
        )
    assert not (tmp_path / "ledger.json").exists()
    assert not (tmp_path / "survey.xlsx").exists()


def test_run_pipeline_records_retrieval(tmp_path):
    cfg = _cfg(tmp_path)
    result = run_pipeline(
        cfg, registry={"fake": FakeSource()}, scorer=FakeScorer(), resolve_bibtex=False
    )
    assert [qr.source for qr in result.ledger.retrieval] == ["fake"]
    assert result.ledger.retrieval[0].retrieved == 3
    led = load_ledger(tmp_path / "ledger.json")
    assert led.retrieval[0].source == "fake"


def test_run_pipeline_snowball_inline(tmp_path):
    from surveyer.config import SnowballConfig
    from surveyer.models import Record
    from surveyer.snowball import SnowballFetch

    cfg = _cfg(tmp_path)
    cfg.snowball = SnowballConfig(enabled=True, direction="both")

    class FakeSnowball:
        def fetch(self, seeds, sb_cfg, *, cancel=None):
            # one fresh, keyword+llm-passing candidate
            return SnowballFetch(
                candidates=[
                    Record(title="Relevant security via citation", abstract="relevant")
                ],
                seeds_total=len(seeds),
                seeds_resolved=len(seeds),
                backward=1,
                forward=0,
                retrieval=[],
            )

    result = run_pipeline(
        cfg,
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolve_bibtex=False,
        snowball=FakeSnowball(),
    )
    assert result.ledger.included == 1  # main arm A unchanged
    assert result.ledger.snowball is not None
    assert result.ledger.snowball.included == 1  # set B
    assert result.ledger.total_included() == 2
    # the snowball include is in the exported kept set
    assert any("via citation" in r.title for r in result.kept)
    led = load_ledger(tmp_path / "ledger.json")
    assert led.snowball is not None and led.snowball.included == 1


def test_run_pipeline_snowball_disabled_by_default(tmp_path):
    result = run_pipeline(
        _cfg(tmp_path),
        registry={"fake": FakeSource()},
        scorer=FakeScorer(),
        resolve_bibtex=False,
    )
    assert result.ledger.snowball is None
