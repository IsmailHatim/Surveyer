from __future__ import annotations

import polars as pl
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from surveyer.export import (
    _to_frame,
    export_bibtex,
    export_csv,
    export_extended,
    export_extended_xlsx,
    export_results,
    export_xlsx,
)
from surveyer.models import Ledger, QueryRetrieval, Record, SourceCount


def test_export_creates_sheets(tmp_path):
    kept = [Record(title="Kept paper", doi="10.1/a", sources=["dblp"], llm_score=0.8)]
    excluded = [Record(title="Dropped paper", sources=["openalex"])]
    ledger = Ledger(identified=[SourceCount(source="dblp", count=1)], included=1)
    out = tmp_path / "survey.xlsx"
    export_xlsx(kept, excluded, ledger, out)

    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"papers", "excluded", "summary", "retrieval"}
    assert wb["papers"]["A1"].value == "title"
    assert wb["papers"]["A2"].value == "Kept paper"
    assert wb["excluded"]["A1"].value == "title"
    assert wb["summary"]["A1"].value == "stage"
    assert wb["summary"]["B1"].value == "count"


def test_export_xlsx_pins_row_heights(tmp_path):
    # Multiline bibtex cells must not autofit rows to a full BibTeX entry
    kept = [
        Record(title="P", bibtex="@article{k,\n  title = {T}\n}", bibtex_source="dblp")
    ]
    out = tmp_path / "survey.xlsx"
    export_xlsx(kept, [], Ledger(), out)

    row = load_workbook(out)["papers"].row_dimensions[2]
    assert row.customHeight, "bibtex row height not pinned; rows will auto-expand"


def test_export_handles_empty_inputs(tmp_path):
    from surveyer.models import Ledger

    out = tmp_path / "empty.xlsx"
    export_xlsx([], [], Ledger(), out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"papers", "excluded", "summary", "retrieval"}
    assert wb["papers"]["A1"].value == "title"


def test_export_csv_writes_three_files(tmp_path):
    kept = [Record(title="Kept paper", doi="10.1/a", sources=["dblp"], llm_score=0.8)]
    excluded = [Record(title="Dropped paper", sources=["openalex"])]
    ledger = Ledger(identified=[SourceCount(source="dblp", count=1)], included=1)

    export_csv(kept, excluded, ledger, tmp_path)

    papers = pl.read_csv(tmp_path / "papers.csv")
    excluded_df = pl.read_csv(tmp_path / "excluded.csv")
    summary = pl.read_csv(tmp_path / "summary.csv")
    assert papers["title"].to_list() == ["Kept paper"]
    assert papers["llm_score"].to_list() == [0.8]
    assert excluded_df["title"].to_list() == ["Dropped paper"]
    assert summary.columns == ["stage", "count"]


def test_export_csv_handles_empty_inputs(tmp_path):
    export_csv([], [], Ledger(), tmp_path)

    papers = pl.read_csv(tmp_path / "papers.csv")
    assert papers.columns[0] == "title"
    assert papers.height == 0


def test_export_results_dispatches_xlsx(tmp_path):
    kept = [Record(title="Kept paper", sources=["dblp"], llm_score=0.8)]
    export_results(kept, [], Ledger(included=1), tmp_path, fmt="xlsx")

    assert (tmp_path / "survey.xlsx").exists()
    wb = load_workbook(tmp_path / "survey.xlsx")
    assert set(wb.sheetnames) == {"papers", "excluded", "summary", "retrieval"}


def test_export_results_dispatches_csv(tmp_path):
    kept = [Record(title="Kept paper", sources=["dblp"], llm_score=0.8)]
    export_results(kept, [], Ledger(included=1), tmp_path, fmt="csv")

    assert not (tmp_path / "survey.xlsx").exists()
    assert {p.name for p in tmp_path.glob("*.csv")} == {
        "papers.csv",
        "excluded.csv",
        "summary.csv",
        "retrieval.csv",
    }


def test_export_results_rejects_unknown_format(tmp_path):
    import pytest

    with pytest.raises(ValueError, match="Unknown export format"):
        export_results([], [], Ledger(), tmp_path, fmt="json")


def test_to_frame_includes_bibtex_columns():
    df = _to_frame([Record(title="P", bibtex="@misc{p}", bibtex_source="local")])
    assert "bibtex" in df.columns
    assert "bibtex_source" in df.columns
    assert df["bibtex"][0] == "@misc{p}"
    assert df["bibtex_source"][0] == "local"


def test_export_includes_exclusion_reason_column(tmp_path):
    excluded = [Record(title="Dropped paper", exclusion_reason="no graph")]
    out = tmp_path / "survey.xlsx"
    export_xlsx([], excluded, Ledger(), out)

    ws = load_workbook(out)["excluded"]
    header = [c.value for c in ws[1]]
    assert "exclusion_reason" in header
    col = header.index("exclusion_reason") + 1
    assert ws.cell(row=2, column=col).value == "no graph"


def test_export_bibtex_writes_references_file(tmp_path):
    kept = [
        Record(title="A", bibtex="@article{a, title={A}}"),
        Record(title="B", bibtex="@misc{b, title={B}}"),
    ]
    export_bibtex(kept, tmp_path)
    text = (tmp_path / "references.bib").read_text()
    assert "@article{a, title={A}}" in text
    assert "@misc{b, title={B}}" in text


def test_export_results_csv_writes_references_and_columns(tmp_path):
    kept = [Record(title="A", bibtex="@misc{a}", bibtex_source="local")]
    export_results(kept, [], Ledger(), tmp_path, fmt="csv")
    assert (tmp_path / "references.bib").exists()
    papers = (tmp_path / "papers.csv").read_text()
    assert "bibtex" in papers
    assert "@misc{a}" in papers


def test_summary_includes_extend_stages_when_extending(tmp_path):
    ledger = Ledger(included=2, previously_included=5, already_screened=3)
    out = tmp_path / "survey.xlsx"
    export_xlsx([], [], ledger, out)

    ws = load_workbook(out)["summary"]
    stages = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "already_screened" in stages
    assert "previously_included" in stages
    assert "total_included" in stages


def test_summary_omits_extend_stages_for_normal_runs(tmp_path):
    out = tmp_path / "survey.xlsx"
    export_xlsx([], [], Ledger(included=2), out)

    ws = load_workbook(out)["summary"]
    stages = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert "previously_included" not in stages


def _screened_baseline(path):
    """A v1 workbook with manual screening artifacts: a fill and a comment."""
    kept = [Record(title="Old kept paper", doi="10.1/old")]
    excluded = [Record(title="Old excluded paper", exclusion_reason="manual")]
    export_xlsx(kept, excluded, Ledger(included=1), path)
    wb = load_workbook(path)
    cell = wb["papers"]["A2"]
    cell.fill = PatternFill(
        start_color="FFFFC000", end_color="FFFFC000", fill_type="solid"
    )
    cell.comment = Comment("double-checked, keep", "screener")
    wb.save(path)


def test_export_extended_xlsx_appends_and_preserves_formatting(tmp_path):
    baseline = tmp_path / "v1.xlsx"
    _screened_baseline(baseline)
    out = tmp_path / "v2" / "survey.xlsx"
    ledger = Ledger(included=1, previously_included=1, already_screened=2)

    export_extended_xlsx(
        baseline,
        [Record(title="Old kept paper", doi="10.1/old")],
        [Record(title="New kept paper", doi="10.1/new", llm_score=0.8)],
        [Record(title="New excluded paper", exclusion_reason="no graph")],
        ledger,
        out,
    )

    wb = load_workbook(out)
    papers = wb["papers"]
    assert papers["A2"].value == "Old kept paper"
    assert papers["A2"].fill.start_color.rgb == "FFFFC000"
    assert papers["A2"].comment is not None
    assert papers["A2"].comment.text == "double-checked, keep"
    assert papers["A3"].value == "New kept paper"

    excluded = wb["excluded"]
    assert excluded["A2"].value == "Old excluded paper"
    assert excluded["A3"].value == "New excluded paper"

    stages = [row[0].value for row in wb["summary"].iter_rows(min_row=2)]
    assert "total_included" in stages


def test_export_extended_xlsx_leaves_baseline_untouched(tmp_path):
    baseline = tmp_path / "v1.xlsx"
    _screened_baseline(baseline)
    before = baseline.read_bytes()

    export_extended_xlsx(
        baseline,
        [Record(title="Old kept paper", doi="10.1/old")],
        [Record(title="New kept paper")],
        [],
        Ledger(included=1, previously_included=1),
        tmp_path / "v2" / "survey.xlsx",
    )

    assert baseline.read_bytes() == before


def test_export_extended_writes_bibtex_for_all_kept(tmp_path):
    baseline = tmp_path / "v1.xlsx"
    _screened_baseline(baseline)
    all_kept = [
        Record(title="Old kept paper", bibtex="@misc{old}"),
        Record(title="New kept paper", bibtex="@misc{new}"),
    ]

    export_extended(
        baseline,
        all_kept,
        [Record(title="New kept paper", bibtex="@misc{new}")],
        [],
        Ledger(included=1, previously_included=1),
        tmp_path / "v2",
    )

    assert (tmp_path / "v2" / "survey.xlsx").exists()
    text = (tmp_path / "v2" / "references.bib").read_text()
    assert "@misc{old}" in text
    assert "@misc{new}" in text


def test_export_extended_backfills_bibtex_on_baseline_rows(tmp_path):
    baseline = tmp_path / "v1.xlsx"
    # Hand-added baseline paper: no bibtex cell in the workbook.
    export_xlsx([Record(title="Hand-added paper")], [], Ledger(included=1), baseline)
    # The pipeline later resolves bibtex on the in-memory record.
    resolved = Record(
        title="Hand-added paper", bibtex="@misc{hand}", bibtex_source="local"
    )
    out = tmp_path / "v2" / "survey.xlsx"

    export_extended_xlsx(
        baseline, [resolved], [], [], Ledger(included=0, previously_included=1), out
    )

    ws = load_workbook(out)["papers"]
    header = [c.value for c in ws[1]]
    bib_col = header.index("bibtex") + 1
    src_col = header.index("bibtex_source") + 1
    assert ws.cell(row=2, column=bib_col).value == "@misc{hand}"
    assert ws.cell(row=2, column=src_col).value == "local"


def test_export_extended_backfills_bibtex_with_padded_title(tmp_path):
    """Baseline title with stray whitespace still backfills from a stripped record."""
    baseline = tmp_path / "v1.xlsx"
    export_xlsx([Record(title="Hand-added paper")], [], Ledger(included=1), baseline)
    # Simulate stray whitespace in cell that was hand-typed.
    wb = load_workbook(baseline)
    ws = wb["papers"]
    header = [c.value for c in ws[1]]
    title_col = header.index("title") + 1
    ws.cell(row=2, column=title_col).value = "  Hand-added paper  "
    wb.save(baseline)

    resolved = Record(
        title="Hand-added paper", bibtex="@misc{hand}", bibtex_source="local"
    )
    out = tmp_path / "v2" / "survey.xlsx"
    export_extended_xlsx(
        baseline, [resolved], [], [], Ledger(included=0, previously_included=1), out
    )

    ws2 = load_workbook(out)["papers"]
    header2 = [c.value for c in ws2[1]]
    bib_col = header2.index("bibtex") + 1
    assert ws2.cell(row=2, column=bib_col).value == "@misc{hand}"


def test_export_extended_backfills_bibtex_by_doi_after_title_edit(tmp_path):
    """A manually edited title still backfills bibtex when the DOI matches."""
    baseline = tmp_path / "v1.xlsx"
    export_xlsx(
        [Record(title="Original title", doi="10.1/x")],
        [],
        Ledger(included=1),
        baseline,
    )
    # The reviewer rewords the title in the screened workbook; the DOI stays.
    wb = load_workbook(baseline)
    ws = wb["papers"]
    header = [c.value for c in ws[1]]
    title_col = header.index("title") + 1
    ws.cell(row=2, column=title_col).value = "A Better, Edited Title"
    wb.save(baseline)

    resolved = Record(
        title="Original title",
        doi="10.1/x",
        bibtex="@article{x}",
        bibtex_source="dblp",
    )
    out = tmp_path / "v2" / "survey.xlsx"
    export_extended_xlsx(
        baseline, [resolved], [], [], Ledger(included=0, previously_included=1), out
    )

    ws2 = load_workbook(out)["papers"]
    header2 = [c.value for c in ws2[1]]
    bib_col = header2.index("bibtex") + 1
    src_col = header2.index("bibtex_source") + 1
    assert ws2.cell(row=2, column=bib_col).value == "@article{x}"
    assert ws2.cell(row=2, column=src_col).value == "dblp"


def _retrieval_ledger():
    return Ledger(
        included=1,
        retrieval=[
            QueryRetrieval(source="openalex", query_label="q1", requested=100,
                           retrieved=100, api_total=5000),
            QueryRetrieval(source="dblp", query_label="q1", requested=100,
                           retrieved=7, api_total=7),
        ],
    )


def test_export_xlsx_has_retrieval_sheet(tmp_path):
    from openpyxl import load_workbook

    path = tmp_path / "survey.xlsx"
    export_xlsx([], [], _retrieval_ledger(), path)
    wb = load_workbook(path)
    assert "retrieval" in wb.sheetnames
    header = [c.value for c in wb["retrieval"][1]]
    assert header == [
        "source", "query_label", "requested", "retrieved", "api_total", "truncated"
    ]
    first = [c.value for c in wb["retrieval"][2]]
    assert first[0] == "openalex"
    assert first[5] is True  # truncated (100 < 5000)


def test_export_csv_writes_retrieval(tmp_path):
    export_csv([], [], _retrieval_ledger(), tmp_path)
    text = (tmp_path / "retrieval.csv").read_text()
    assert "source,query_label,requested,retrieved,api_total,truncated" in text
    assert "openalex" in text
