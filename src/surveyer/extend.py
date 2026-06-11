"""Load a manually screened survey workbook and reconcile it with a new run."""

from __future__ import annotations

from pathlib import Path

import msgspec
from openpyxl import load_workbook
from rapidfuzz import fuzz

from surveyer.dedup import _is_weak_doi, normalize_title
from surveyer.models import Record

_REQUIRED_SHEETS = ("papers", "excluded")
_TITLE_THRESHOLD = 90


class Baseline(msgspec.Struct):
    """Manually screened records: pinned includes and final excludes."""

    included: list[Record]
    excluded: list[Record]


def _str_or_none(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _split_list(value: object) -> list[str]:
    text = _str_or_none(value)
    if text is None:
        return []
    return [part.strip() for part in text.split(";") if part.strip()]


def _int_or_none(value: object) -> int | None:
    text = _str_or_none(value)
    return int(float(text)) if text is not None else None


def _float_or_none(value: object) -> float | None:
    text = _str_or_none(value)
    return float(text) if text is not None else None


def _records_from_sheet(ws, sheet_name: str) -> list[Record]:
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return []
    idx = {str(name): i for i, name in enumerate(header) if name is not None}
    if "title" not in idx:
        raise ValueError(f"Sheet {sheet_name!r} has no 'title' column.")

    records: list[Record] = []
    for row_num, row in enumerate(rows, start=2):
        if all(v is None or not str(v).strip() for v in row):
            continue  # blank padding row

        def cell(column: str, row: tuple = row) -> object:
            i = idx.get(column)
            return row[i] if i is not None and i < len(row) else None

        def number(column: str, parse):  # type: ignore[no-untyped-def]
            value = cell(column)
            try:
                return parse(value)
            except ValueError as exc:
                raise ValueError(
                    f"Sheet {sheet_name!r} row {row_num} has an invalid "
                    f"number in column {column!r}: {value!r}"
                ) from exc

        title = _str_or_none(cell("title"))
        if title is None:
            raise ValueError(f"Sheet {sheet_name!r} row {row_num} has an empty title.")
        records.append(
            Record(
                title=title,
                doi=_str_or_none(cell("doi")),
                authors=_split_list(cell("authors")),
                year=number("year", _int_or_none),
                venue=_str_or_none(cell("venue")),
                abstract=_str_or_none(cell("abstract")),
                keywords=_split_list(cell("keywords")),
                url=_str_or_none(cell("url")),
                n_citations=number("n_citations", _int_or_none),
                sources=_split_list(cell("sources")),
                query_labels=_split_list(cell("query_labels")),
                llm_score=number("llm_score", _float_or_none),
                llm_reason=_str_or_none(cell("llm_reason")),
                exclusion_reason=_str_or_none(cell("exclusion_reason")),
                bibtex=_str_or_none(cell("bibtex")),
                bibtex_source=_str_or_none(cell("bibtex_source")),
            )
        )
    return records


def load_baseline(path: str | Path) -> Baseline:
    """Parse the papers excluded sheets of a screened workbook into records."""
    path = Path(path)
    if not path.is_file():
        raise ValueError(f"Baseline xlsx not found: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        missing = [s for s in _REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            raise ValueError(
                f"Baseline xlsx {path} is missing sheet(s): {', '.join(missing)}"
            )
        return Baseline(
            included=_records_from_sheet(wb["papers"], "papers"),
            excluded=_records_from_sheet(wb["excluded"], "excluded"),
        )
    finally:
        wb.close()


def split_already_screened(
    records: list[Record], baseline: Baseline
) -> tuple[list[Record], int]:
    """Partition new records into (fresh, n_dropped) against screened papers."""
    screened = baseline.included + baseline.excluded
    known_dois = {r.doi.lower().strip() for r in screened if r.doi}
    # Pair normalized title with normalized DOI for the fuzzy pass.
    known_title_doi: list[tuple[str, str | None]] = [
        (normalize_title(r.title), r.doi.lower().strip() if r.doi else None)
        for r in screened
    ]

    fresh: list[Record] = []
    dropped = 0
    for r in records:
        doi = r.doi.lower().strip() if r.doi else None
        if doi and doi in known_dois:
            dropped += 1
            continue
        norm = normalize_title(r.title)
        fuzzy_match = False
        for ex_title, ex_doi in known_title_doi:
            # Conflicting strong DOIs on both sides block the fuzzy drop.
            if (
                doi
                and ex_doi
                and ex_doi != doi
                and not (_is_weak_doi(doi) or _is_weak_doi(ex_doi))
            ):
                continue
            if fuzz.token_sort_ratio(norm, ex_title) >= _TITLE_THRESHOLD:
                fuzzy_match = True
                break
        if fuzzy_match:
            dropped += 1
            continue
        fresh.append(r)
    return fresh, dropped
