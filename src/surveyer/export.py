"""Export results to an Excel sheet."""

from __future__ import annotations

import copy
import shutil
from pathlib import Path
from typing import TYPE_CHECKING

import polars as pl
import xlsxwriter
from openpyxl import load_workbook

from surveyer.dedup import normalize_title
from surveyer.models import Ledger, Record

if TYPE_CHECKING:
    from polars._typing import (
        ColumnWidthsDefinition,
        ConditionalFormatDict,
        SchemaDict,
    )

# Explicit schema so polars never infers a column type from a leading run of
# null values (e.g. records with no llm_score) and then chokes on a later float.
_SCHEMA: SchemaDict = {
    "title": pl.Utf8,
    "authors": pl.Utf8,
    "year": pl.Int64,
    "venue": pl.Utf8,
    "doi": pl.Utf8,
    "url": pl.Utf8,
    "abstract": pl.Utf8,
    "keywords": pl.Utf8,
    "n_citations": pl.Int64,
    "sources": pl.Utf8,
    "query_labels": pl.Utf8,
    "llm_score": pl.Float64,
    "llm_reason": pl.Utf8,
    "exclusion_reason": pl.Utf8,
    "bibtex": pl.Utf8,
    "bibtex_source": pl.Utf8,
}

_COLUMNS = list(_SCHEMA)

# Bold, shaded header row applied to every sheet.
_HEADER_FORMAT = {
    "bold": True,
    "bg_color": "#D9E1F2",
    "border": 1,
    "align": "center",
    "valign": "vcenter",
}

# Fixed column widths (in pixels, as polars expects) instead of autofit, so
# every column keeps a predictable size and the sheet fits on one screen.
_COLUMN_WIDTHS: ColumnWidthsDefinition = {
    "title": 290,
    "authors": 200,
    "year": 50,
    "venue": 130,
    "doi": 150,
    "url": 190,
    "abstract": 360,
    "keywords": 170,
    "n_citations": 70,
    "sources": 100,
    "query_labels": 110,
    "llm_score": 75,
    "llm_reason": 290,
    "exclusion_reason": 180,
    "bibtex": 360,
    "bibtex_source": 90,
}

# Yellow (low) to green (high) colour scale over the 0..1 llm_score.
_SCORE_COLOR_SCALE: ConditionalFormatDict = {
    "llm_score": {
        "type": "2_color_scale",
        "min_type": "num",
        "min_value": 0.0,
        "min_color": "#FFEB84",
        "max_type": "num",
        "max_value": 1.0,
        "max_color": "#63BE7B",
    }
}

# Single-line rows excel default size
_ROW_HEIGHT = 21

# Highlight locally-synthesized entries in red.
_BIBTEX_FLAG: ConditionalFormatDict = {
    "bibtex_source": {
        "type": "text",
        "criteria": "containing",
        "value": "local",
        "format": {"bg_color": "#FFC7CE", "font_color": "#9C0006"},
    }
}


def _record_row(r: Record) -> dict:
    """Map one record onto the export schema's column names."""
    return {
        "title": r.title,
        "authors": "; ".join(r.authors),
        "year": r.year,
        "venue": r.venue,
        "doi": r.doi,
        "url": r.url,
        "abstract": r.abstract,
        "keywords": "; ".join(r.keywords),
        "n_citations": r.n_citations,
        "sources": "; ".join(r.sources),
        "query_labels": "; ".join(r.query_labels),
        "llm_score": r.llm_score,
        "llm_reason": r.llm_reason,
        "exclusion_reason": r.exclusion_reason,
        "bibtex": r.bibtex,
        "bibtex_source": r.bibtex_source,
    }


def _to_frame(records: list[Record]) -> pl.DataFrame:
    rows = [_record_row(r) for r in records]
    return pl.DataFrame(rows, schema=_SCHEMA).select(_COLUMNS)


def _retrieval_frame(ledger: Ledger) -> pl.DataFrame:
    """Per-(source, query) retrieval detail: requested/retrieved/api_total/truncated.

    `truncated` is computed per query here; PRISMA aggregates it per source.
    """
    schema = {
        "source": pl.Utf8,
        "query_label": pl.Utf8,
        "requested": pl.Int64,
        "retrieved": pl.Int64,
        "api_total": pl.Int64,
        "truncated": pl.Boolean,
    }
    rows = [
        {
            "source": qr.source,
            "query_label": qr.query_label,
            "requested": qr.requested,
            "retrieved": qr.retrieved,
            "api_total": qr.api_total,
            "truncated": qr.api_total is not None and qr.retrieved < qr.api_total,
        }
        for qr in ledger.retrieval
        + (ledger.snowball.retrieval if ledger.snowball is not None else [])
    ]
    return pl.DataFrame(rows, schema=schema)


def _summary_frame(ledger: Ledger) -> pl.DataFrame:
    stages: list[tuple[str, int]] = [
        ("total_identified", ledger.total_identified()),
        ("duplicates_removed", ledger.duplicates_removed),
        ("after_dedup", ledger.after_dedup()),
        ("excluded_keyword", ledger.excluded_keyword),
        ("excluded_llm", ledger.excluded_llm),
        ("included", ledger.included),
    ]
    if ledger.snowball is not None:
        sb = ledger.snowball
        stages += [
            ("snowball_identified", sb.identified),
            ("snowball_duplicates_removed", sb.duplicates_removed),
            ("snowball_excluded_keyword", sb.excluded_keyword),
            ("snowball_excluded_llm", sb.excluded_llm),
            ("snowball_included", sb.included),
        ]
    if (
        ledger.previously_included
        or ledger.already_screened
        or ledger.snowball is not None
    ):
        stages += [
            ("already_screened", ledger.already_screened),
            ("previously_included", ledger.previously_included),
            ("total_included", ledger.total_included()),
        ]
    return pl.DataFrame(
        {"stage": [s for s, _ in stages], "count": [c for _, c in stages]}
    )


def export_bibtex(kept: list[Record], out_dir: str | Path) -> None:
    """Write kept records BibTeX entries to references.bib in out_dir."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    entries = [r.bibtex for r in kept if r.bibtex]
    text = "\n\n".join(entries)
    (out_dir / "references.bib").write_text(text + "\n" if text else "")


def export_xlsx(
    kept: list[Record],
    excluded: list[Record],
    ledger: Ledger,
    path: str | Path,
) -> None:
    """Write kept and excluded records and a ledger summary to a .xlsx."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = xlsxwriter.Workbook(str(path))
    try:
        # Deep-copy the combined conditional formats on each call
        _to_frame(kept).write_excel(
            workbook=wb,
            worksheet="papers",
            header_format=_HEADER_FORMAT,
            conditional_formats=copy.deepcopy({**_SCORE_COLOR_SCALE, **_BIBTEX_FLAG}),
            column_widths=_COLUMN_WIDTHS,
            row_heights=_ROW_HEIGHT,
            freeze_panes=(1, 0),
        )
        _to_frame(excluded).write_excel(
            workbook=wb,
            worksheet="excluded",
            header_format=_HEADER_FORMAT,
            conditional_formats=copy.deepcopy({**_SCORE_COLOR_SCALE, **_BIBTEX_FLAG}),
            column_widths=_COLUMN_WIDTHS,
            row_heights=_ROW_HEIGHT,
            freeze_panes=(1, 0),
        )
        _summary_frame(ledger).write_excel(
            workbook=wb,
            worksheet="summary",
            header_format=_HEADER_FORMAT,
            autofit=True,
            freeze_panes=(1, 0),
        )
        _retrieval_frame(ledger).write_excel(
            workbook=wb,
            worksheet="retrieval",
            header_format=_HEADER_FORMAT,
            autofit=True,
            freeze_panes=(1, 0),
        )
    finally:
        wb.close()


def export_csv(
    kept: list[Record],
    excluded: list[Record],
    ledger: Ledger,
    out_dir: str | Path,
) -> None:
    """Write kept and excluded records and a ledger summary as CSV files."""
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    _to_frame(kept).write_csv(out_dir / "papers.csv")
    _to_frame(excluded).write_csv(out_dir / "excluded.csv")
    _summary_frame(ledger).write_csv(out_dir / "summary.csv")
    _retrieval_frame(ledger).write_csv(out_dir / "retrieval.csv")


def export_results(
    kept: list[Record],
    excluded: list[Record],
    ledger: Ledger,
    out_dir: str | Path,
    fmt: str = "xlsx",
) -> None:
    """Export results to out_dir in the requested format ("xlsx" or "csv")."""
    out_dir = Path(out_dir)
    if fmt == "csv":
        export_csv(kept, excluded, ledger, out_dir)
    elif fmt == "xlsx":
        export_xlsx(kept, excluded, ledger, out_dir / "survey.xlsx")
    else:
        raise ValueError(f"Unknown export format: {fmt!r}")
    export_bibtex(kept, out_dir)


def _append_records(ws, records: list[Record]) -> None:
    """Append records below existing rows, matching columns by header name."""
    header = [cell.value for cell in ws[1]]
    for r in records:
        row = _record_row(r)
        ws.append([row.get(name) for name in header])


def _rewrite_summary(wb, ledger: Ledger) -> None:
    """Replace the summary sheet with the extended ledger counts."""
    if "summary" in wb.sheetnames:
        del wb["summary"]
    ws = wb.create_sheet("summary")
    frame = _summary_frame(ledger)
    ws.append(frame.columns)
    for stage, count in frame.iter_rows():
        ws.append([stage, count])


def _rewrite_retrieval(wb, ledger: Ledger) -> None:
    """Replace the retrieval sheet with the new run's per-query detail."""
    if "retrieval" in wb.sheetnames:
        del wb["retrieval"]
    ws = wb.create_sheet("retrieval")
    frame = _retrieval_frame(ledger)
    ws.append(frame.columns)
    for row in frame.iter_rows():
        ws.append(list(row))


def _patch_missing_bibtex(ws, records: list[Record]) -> None:
    """Backfill empty bibtex/bibtex_source cells, matching by DOI then title.

    DOI is preferred because the extend workflow lets reviewers hand-edit titles
    in the screened workbook; a normalized-title match is the fallback.
    """
    header = [cell.value for cell in ws[1]]
    if "bibtex" not in header or "title" not in header:
        return
    title_col = header.index("title") + 1
    bib_col = header.index("bibtex") + 1
    src_col = header.index("bibtex_source") + 1 if "bibtex_source" in header else None
    doi_col = header.index("doi") + 1 if "doi" in header else None

    # Index resolved records by normalized DOI (preferred) and normalized title.
    by_doi: dict[str, Record] = {}
    by_title: dict[str, Record] = {}
    for r in records:
        if not r.bibtex:
            continue
        if r.doi and r.doi.strip():
            by_doi.setdefault(r.doi.lower().strip(), r)
        norm = normalize_title(r.title)
        if norm:
            by_title.setdefault(norm, r)

    for row in ws.iter_rows(min_row=2):
        bib_cell = row[bib_col - 1]
        if bib_cell.value:
            continue
        record = None
        if doi_col is not None:
            doi_val = row[doi_col - 1].value
            if doi_val is not None and str(doi_val).strip():
                record = by_doi.get(str(doi_val).lower().strip())
        if record is None:
            title_val = row[title_col - 1].value
            norm = normalize_title(str(title_val)) if title_val is not None else ""
            if norm:
                record = by_title.get(norm)
        if record is None:
            continue
        bib_cell.value = record.bibtex
        if src_col is not None:
            row[src_col - 1].value = record.bibtex_source


def export_extended_xlsx(
    baseline_path: str | Path,
    all_kept: list[Record],
    new_kept: list[Record],
    new_excluded: list[Record],
    ledger: Ledger,
    path: str | Path,
) -> None:
    """Copy the screened workbook to path and append the new run's records."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(baseline_path, path)
    wb = load_workbook(path)
    _append_records(wb["papers"], new_kept)
    _append_records(wb["excluded"], new_excluded)
    _patch_missing_bibtex(wb["papers"], all_kept)
    _rewrite_summary(wb, ledger)
    _rewrite_retrieval(wb, ledger)
    wb.save(path)


def export_extended(
    baseline_path: str | Path,
    all_kept: list[Record],
    new_kept: list[Record],
    new_excluded: list[Record],
    ledger: Ledger,
    out_dir: str | Path,
) -> None:
    """Write the extended survey.xlsx plus references.bib for all included."""
    out_dir = Path(out_dir)
    export_extended_xlsx(
        baseline_path, all_kept, new_kept, new_excluded, ledger, out_dir / "survey.xlsx"
    )
    export_bibtex(all_kept, out_dir)
